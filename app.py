from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from account_lookup import load_accounts
from excel_service import build_excel
from xml_service import (
    build_employee_summary,
    extract_contributions,
    extract_fund_deposits,
    load_xsd,
    to_bytes,
    validate_and_repair,
    validate_xsd,
)

import re


def is_valid_dat_filename(filename: str) -> bool:
    """
    בודק אם שם הקובץ כבר בפורמט התקין של ממשק המעסיקים.
    דוגמה:
    006000058893207EMPONG000006202608130927000001.DAT
    """

    pattern = (
        r"^006000"
        r"\d{9}"
        r"EMPONG"
        r"000006"
        r"\d{14}"
        r"\d{4}"
        r"\.DAT$"
    )

    return re.fullmatch(pattern, filename.upper()) is not None


def build_correct_dat_filename(tree) -> str:
    """
    יוצר שם קובץ DAT תקין לפי הנתונים שבתוך ה-XML.
    """

    root = tree.getroot()

    def get_text(tag: str) -> str:
        nodes = root.xpath(
            f".//*[local-name()='{tag}']"
        )

        if not nodes:
            return ""

        return (nodes[0].text or "").strip()

    sender_id = "".join(
        ch
        for ch in get_text("MISPAR-ZIHUI-SHOLECH")
        if ch.isdigit()
    )

    version = get_text("MISPAR-GIRSAT-XML") or "006"
    file_number = get_text("MISPAR-HAKOVETZ")

    if not sender_id:
        sender_id = "058893207"

    sender_id = sender_id.zfill(9)

    if len(file_number) < 14:
        raise ValueError(
            "לא ניתן ליצור שם קובץ תקין: MISPAR-HAKOVETZ אינו תקין."
        )

    timestamp = file_number[:14]
    serial = file_number[-4:] if len(file_number) >= 4 else "0001"

    return (
        f"006000"
        f"{sender_id}"
        f"EMPONG"
        f"000{version}"
        f"{timestamp}"
        f"{serial}"
        f".DAT"
    )

# ============================================================
# פונקציה ליצירת Excel עם גיליון אחד בלבד
# ============================================================

def extract_single_sheet_excel(
    excel_bytes: bytes,
    sheet_name: str,
) -> bytes | None:
    """
    יוצר קובץ Excel חדש שמכיל רק גיליון אחד
    מתוך קובץ Excel קיים.

    אם אין התאמה מדויקת בשם,
    מנסה לזהות אוטומטית גיליון סיכום קופות/הפקדות.
    """

    source = BytesIO(excel_bytes)
    workbook = load_workbook(source)

    target_sheet = None

    # 1. התאמה מדויקת
    if sheet_name in workbook.sheetnames:
        target_sheet = sheet_name

    # 2. שמות אפשריים נוספים
    if target_sheet is None:
        possible_names = [
            "סיכום קופות והפקדות",
            "סיכום קופות",
            "קופות והפקדות",
            "סיכום הפקדות",
            "סיכום פקודות והפקדות",
        ]

        for possible_name in possible_names:
            if possible_name in workbook.sheetnames:
                target_sheet = possible_name
                break

    # 3. חיפוש לפי מילות מפתח
    if target_sheet is None:
        for current_sheet in workbook.sheetnames:
            if (
                "קופות" in current_sheet
                or "הפקדות" in current_sheet
            ):
                target_sheet = current_sheet
                break

    # אם עדיין לא נמצא גיליון
    if target_sheet is None:
        return None

    # מוחקים את כל הגיליונות האחרים
    for current_sheet in workbook.sheetnames.copy():
        if current_sheet != target_sheet:
            del workbook[current_sheet]

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()


# ============================================================
# הגדרות
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

XSD_PATH = (
    BASE_DIR
    / "mimshak_maasikim_shotef_xsd_schema_006.xsd.xml"
)

ACCOUNTS_PATH = BASE_DIR / "accounts.xlsx"


# ============================================================
# הגדרות Streamlit
# ============================================================

st.set_page_config(
    page_title="LYN XML Fix",
    page_icon="✅",
    layout="wide",
)

st.markdown(
    """
    <style>
    html, body, [class*='css'] {
        direction: rtl;
    }

    [data-testid='stDataFrame'] {
        direction: rtl;
    }

    .block-container {
        padding-top: 2rem;
        max-width: 1250px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# כותרת
# ============================================================

st.title("תיקון קובץ XML והפקת טבלת עובדים")

st.caption(
    "מעלים קובץ XML/DAT ומקבלים "
    "קובץ מתוקן ו-Excel עם העובדים וההפרשות."
)


# ============================================================
# Sidebar
# ============================================================

with st.sidebar:

    st.header("הגדרות")

    employer_option = st.selectbox(
        "סוג מעסיק",
        [
            "ללא שינוי",
            "ח.פ – קוד 1",
            "עוסק מורשה – קוד 5",
        ],
    )

    employer_type = {
        "ח.פ – קוד 1": "1",
        "עוסק מורשה – קוד 5": "5",
    }.get(employer_option)

    last_deposit_option = st.selectbox(
        "הפקדה אחרונה כשחסר השדה",
        [
            "2 – כן",
            "1 – לא",
        ],
    )

    last_deposit = last_deposit_option.split(" ")[0]

    repair_banks = st.checkbox(
        "לתקן שדות בנק שערכם 0",
        value=True,
    )


# ============================================================
# העלאת קובץ
# ============================================================

uploaded = st.file_uploader(
    "בחר קובץ XML או DAT",
    type=[
        "xml",
        "dat",
        "DAT",
    ],
)

if not uploaded:

    st.info(
        "יש להעלות קובץ כדי להתחיל."
    )

    st.stop()


# ============================================================
# טעינת XSD
# ============================================================

try:

    schema = load_xsd(
        XSD_PATH
    )

except Exception as exc:

    st.error(
        f"לא ניתן לטעון את קובץ ה-XSD: {exc}"
    )

    st.stop()


# ============================================================
# טעינת accounts.xlsx
# ============================================================

try:

    accounts = load_accounts(
        ACCOUNTS_PATH
    )

except Exception as exc:

    st.error(
        f"לא ניתן לטעון את קובץ חשבונות הקופות: {exc}"
    )

    st.stop()


# ============================================================
# קריאת הקובץ + תיקון
# ============================================================

try:

    original_tree = __import__(
        "xml_service"
    ).parse_xml(
        uploaded.getvalue()
    )

    original_errors = validate_xsd(
        original_tree,
        schema,
    )

    fixed_tree, changes = validate_and_repair(
        uploaded.getvalue(),
        employer_id_type=employer_type,
        last_deposit_default=last_deposit,
        repair_bank_zeroes=repair_banks,
        accounts=accounts,
    )

    final_errors = validate_xsd(
        fixed_tree,
        schema,
    )

except Exception as exc:

    st.error(
        f"לא ניתן לקרוא או לעבד את הקובץ: {exc}"
    )

    st.stop()


# ============================================================
# נתוני עובדים והפרשות
# ============================================================

contribution_rows = extract_contributions(
    fixed_tree
)

employee_rows = build_employee_summary(
    contribution_rows
)

change_rows = [
    change.to_dict()
    for change in changes
]


# ============================================================
# סיכום קופות והפקדות
# ============================================================

fund_rows = extract_fund_deposits(
    fixed_tree,
    accounts,
)


# ============================================================
# מספר עובדים ייחודיים
# ============================================================

unique_employee_ids = {
    str(
        row.get(
            "ת.ז",
            "",
        )
    ).strip()
    for row in employee_rows
    if str(
        row.get(
            "ת.ז",
            "",
        )
    ).strip()
}

employee_count = len(
    unique_employee_ids
)


# ============================================================
# מספר שורות דיווח
# ============================================================

report_rows_count = len(
    fixed_tree.xpath(
        "//*[local-name()='ChodeshMaskoretVestatusOved']"
    )
)


# ============================================================
# שם מעסיק
# ============================================================

employer_name_nodes = fixed_tree.xpath(
    "//*[local-name()='SHEM-MAASIK']"
)

employer_name = ""

if employer_name_nodes:

    employer_name = (
        employer_name_nodes[0].text
        or ""
    ).strip()


if employer_name:

    st.subheader(
        f"מעסיק: {employer_name}"
    )

else:

    st.subheader(
        "מעסיק: לא זוהה"
    )


# ============================================================
# מדדים
# ============================================================

c1, c2, c3 = st.columns(3)

c1.metric(
    "עובדים",
    employee_count,
)

c2.metric(
    "שורות דיווח",
    report_rows_count,
)

c3.metric(
    "תיקונים שבוצעו",
    len(changes),
)


# ============================================================
# סטטוס XSD
# ============================================================

if not final_errors:

    if changes:

        st.success(
            "הקובץ תוקן ועבר בהצלחה "
            "אימות מלא מול XSD גרסה 006."
        )

    else:

        st.success(
            "הקובץ כבר תקין ועבר בהצלחה "
            "אימות מלא מול XSD גרסה 006."
        )

else:

    st.warning(
        "בוצעו תיקונים אוטומטיים, "
        "אך נשארו שגיאות XSD שלא ניתן "
        "לתקן ללא מידע נוסף. "
        "ניתן לראות אותן בלשונית השגיאות."
    )


# ============================================================
# יצירת קבצי הורדה
# ============================================================

fixed_xml = to_bytes(
    fixed_tree
)

excel_bytes = build_excel(
    employee_rows,
    contribution_rows,
    change_rows,
    final_errors,
    fund_rows=fund_rows,
)


# ============================================================
# יצירת Excel של סיכום קופות והפקדות בלבד
# ============================================================

summary_excel_bytes = extract_single_sheet_excel(
    excel_bytes,
    "סיכום קופות והפקדות",
)


# ============================================================
# שם בסיס לקבצים
# ============================================================

base = uploaded.name.rsplit(
    ".",
    1,
)[0]


# ============================================================
# כפתורי הורדה
# ============================================================

b1, b2, b3 = st.columns(3)
  if is_valid_dat_filename(uploaded.name):
    correct_dat_filename = uploaded.name
else:
    correct_dat_filename = build_correct_dat_filename(
        fixed_tree
    )

with b1:


    st.download_button(
        "הורדת XML/DAT מתוקן",
        data=fixed_xml,
        file_name=correct_dat_filename,
        mime="application/xml",
        use_container_width=True,
        type="primary",
    )


with b2:

    st.download_button(
        "הורדת Excel עובדים והפרשות",
        data=excel_bytes,
        file_name=(
            f"{base}_EMPLOYEES_AND_CONTRIBUTIONS.xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
    )


with b3:
    if summary_excel_bytes is not None:

        # שם מעסיק
        employer_nodes = fixed_tree.xpath(
            "//*[local-name()='SHEM-MAASIK']"
        )

        employer_name = (
            (employer_nodes[0].text or "").strip()
            if employer_nodes
            else "מעסיק"
        )

        # חודש דיווח
        month_nodes = fixed_tree.xpath(
            "//*[local-name()='CHODESH-MASKORET']"
        )

        report_month = ""

        if month_nodes:
            month_text = (month_nodes[0].text or "").strip()

            # לדוגמה: 2026-07-01 -> 072026
            if len(month_text) >= 7:
                year = month_text[0:4]
                month = month_text[5:7]
                report_month = f"{month}{year}"

        # ניקוי תווים שאסור שיהיו בשם קובץ ב-Windows
        for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
            employer_name = employer_name.replace(char, "")

        summary_filename = f"{employer_name} {report_month}.xlsx"

        st.download_button(
            "📊 סיכום קופות והפקדות",
            data=summary_excel_bytes,
            file_name=summary_filename,
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
        )

    else:
        st.info("לא נמצא גיליון 'סיכום פקודות והפקדות' בקובץ.")

        st.warning(
            "לא נמצא גיליון סיכום קופות והפקדות "
            "בקובץ ה-Excel."
        )


# ============================================================
# לשוניות
# ============================================================

tab1, tab2, tab3, tab4, tab5 = st.tabs(
    [
        "סיכום קופות והפקדות",
        "סיכום עובדים",
        "פירוט הפרשות",
        "תיקונים",
        "שגיאות XSD",
    ]
)


# ============================================================
# לשונית 1 - קופות והפקדות
# ============================================================

with tab1:

    if fund_rows:

        fund_df = pd.DataFrame(
            fund_rows
        )

        preferred_columns = [
            "שם קופה",
            "סכום להפקדה",
            "פרטי חשבון",
            "מספר קופה",
            "שם חברה מנהלת",
            "ח.פ חברה מנהלת",
            "סטטוס זיהוי",
        ]

        existing_columns = [
            column
            for column in preferred_columns
            if column in fund_df.columns
        ]

        st.dataframe(
            fund_df[
                existing_columns
            ],
            use_container_width=True,
            hide_index=True,
            column_config={
                "סכום להפקדה": (
                    st.column_config.NumberColumn(
                        "סכום להפקדה",
                        format="₪ %.2f",
                    )
                )
            },
        )

    else:

        st.info(
            "לא נמצאו קופות או סכומים להפקדה."
        )


# ============================================================
# לשונית 2 - עובדים
# ============================================================

with tab2:

    if employee_rows:

        st.dataframe(
            pd.DataFrame(
                employee_rows
            ),
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.info(
            "לא נמצאו עובדים או הפרשות בקובץ."
        )


# ============================================================
# לשונית 3 - הפרשות
# ============================================================

with tab3:

    if contribution_rows:

        st.dataframe(
            pd.DataFrame(
                contribution_rows
            ),
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.info(
            "לא נמצאו רשומות הפרשה."
        )


# ============================================================
# לשונית 4 - תיקונים
# ============================================================

with tab4:

    if change_rows:

        st.dataframe(
            pd.DataFrame(
                change_rows
            ),
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.success(
            "לא נדרשו תיקונים אוטומטיים."
        )


# ============================================================
# לשונית 5 - שגיאות XSD
# ============================================================

with tab5:

    st.caption(
        f"לפני התיקון: "
        f"{len(original_errors)} שגיאות | "
        f"אחרי התיקון: "
        f"{len(final_errors)} שגיאות"
    )

    if final_errors:

        st.dataframe(
            pd.DataFrame(
                final_errors
            ),
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.success(
            "אין שגיאות XSD בקובץ הסופי."
        )
